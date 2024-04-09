from selenium import webdriver      
from selenium.webdriver.common.by import By
from openpyxl import Workbook
from openpyxl import Workbook, load_workbook
import time , sys
from datetime import datetime


current_time = datetime.now()

options = webdriver.ChromeOptions()
driver = webdriver.Chrome(options=options)
driver.get("https://opensource-demo.orangehrmlive.com/web/index.php/auth/login")
time.sleep(2)
driver.find_element(By.XPATH, "/html/body/div/div[1]/div/div[1]/div/div[2]/div[2]/form/div[1]/div/div[2]/input").send_keys("Admin")
time.sleep(2)
driver.find_element(By.XPATH, "/html/body/div/div[1]/div/div[1]/div/div[2]/div[2]/form/div[2]/div/div[2]/input").send_keys("admin123")
time.sleep(2)
driver.find_element(By.XPATH, "/html/body/div/div[1]/div/div[1]/div/div[2]/div[2]/form/div[3]/button").click()
time.sleep(2)


 # for admin
driver.find_element(By.XPATH, "/html/body/div/div[1]/div[1]/aside/nav/div[2]/ul/li[1]/a/span").click()
time.sleep(5)
elements = driver.find_elements(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[3]/div/div[2]/div")
time.sleep(5)
data_list = []
for i in range(1, len(elements)+1):
    data = {}
    data['user_name'] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[3]/div/div[2]/div[" + str(i) + "]/div/div[2]").text
    data['user_role'] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[3]/div/div[2]/div[" + str(i) + "]/div/div[3]").text
    data['employee_name'] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[3]/div/div[2]/div[" + str(i) + "]/div/div[4]").text
    data['status'] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[3]/div/div[2]/div[" + str(i) + "]/div/div[5]").text
    data_list.append(data)


#data_list = []
#for i in range(1, 30):
    #data = {}
    #data['user name'] = f"TestName{str(i)}"
    #data['user role'] = f"TestRole{str(i)}"
    #data['employee name'] = f"TestEmpName{str(i)}"
    #data['status'] = f"TestStatus{str(i)}"
    #data_list.append(data)

wb = Workbook()
ws = wb.active
ws_index = ws.max_row + 2

ws['A1'] = 'user_name'
ws['B1'] = 'user_role'
ws['C1'] = 'employee_name'
ws['D1'] = 'status'

for data in data_list:
    
    ws['A'+str(ws_index)] = data['user_name']
    ws['B'+str(ws_index)] = data['user_role']
    ws['C'+str(ws_index)] = data['employee_name']
    ws['D'+str(ws_index)] = data['status']
    ws_index += 1
    
wb.save(f'output/orange_report_{current_time.strftime('%d_%m_%Y_%H_%M')}.xlsx')


#for pim
driver.find_element(By.XPATH, "/html/body/div/div[1]/div[1]/aside/nav/div[2]/ul/li[2]/a/span").click()
time.sleep(2)
elements1 = driver.find_elements(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[3]/div/div[2]/div")
time.sleep(2)

data_list1 = []
for i in range(1, len(elements1)+1):
    data1 = {}
    data1['id'] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[3]/div/div[2]/div[" + str(i) +"]/div/div[2]/div").text
    data1['first_name'] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[3]/div/div[2]/div[" + str(i) +"]/div/div[3]/div").text
    data1['last_name'] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[3]/div/div[2]/div[" + str(i) +"]/div/div[4]/div").text
    data1['job_title'] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[3]/div/div[2]/div[" + str(i) +"]/div/div[5]/div").text
    data_list1.append(data1)
print(len(elements1))

ws1 = wb.create_sheet("PIM_orange")
ws1_index = ws1.max_row + 2

ws1['A1'] = 'id'
ws1['B1'] = 'first_name'
ws1['C1'] = 'last_name'
ws1['D1'] = 'job_title'

for data in data_list1:
    
    ws1['A'+str(ws1_index)] = data1['id']
    ws1['B'+str(ws1_index)] = data1['first_name']
    ws1['C'+str(ws1_index)] = data1['last_name']
    ws1['D'+str(ws1_index)] = data1['job_title']
    ws1_index += 1
wb.save(f'output/orange_report_{current_time.strftime('%d_%m_%Y_%H_%M')}.xlsx')



#for leave
driver.find_element(By.XPATH, '/html/body/div/div[1]/div[1]/aside/nav/div[2]/ul/li[3]/a/span').click()
time.sleep(2)
elements2 = driver.find_elements(By.XPATH, '/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[2]/div/div[2]/div')
time.sleep(2)
data_list2 =[]
for i in range(1, len(elements2)+1):
    data2 = {}
    data2['date'] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[2]/div/div[2]/div[" + str(i) + "]/div/div[2]/div").text
    data2['Employee_name'] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[2]/div/div[2]/div[" + str(i) + "]/div/div[3]/div").text
    data2['leave_type'] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[2]/div/div[2]/div[" + str(i) + "]/div/div[4]/div").text
    data2['leave_balance'] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[2]/div/div[2]/div[" + str(i) + "]/div/div[5]/div").text
    data2['Number_of_days'] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[2]/div/div[2]/div[" + str(i) + "]/div/div[6]/div").text
    data2['status'] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[2]/div/div[2]/div[" + str(i) + "]/div/div[7]/div").text
    data2['comments'] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[2]/div/div[2]/div[" + str(i) + "]/div/div[8]/div").text
    data_list2.append(data2)
print(len(elements2))

ws2 = wb.create_sheet('leave_orange')
ws2_index = ws2.max_row + 2

ws2['A1'] = "date"
ws2['B1'] = "Employee_name"
ws2['C1'] = "leave_type"
ws2['D1'] = "leave_balance"
ws2['E1'] = "Number_of_days"
ws2['F1'] = "status"
ws2['G1'] = 'comments'

for data in data_list2:
    
    ws2['A1' +str(ws2_index)] = data2['date']
    ws2['B1' +str(ws2_index)] = data2['Employee_name']
    ws2['C1' +str(ws2_index)] = data2['leave_type']
    ws2['D1' +str(ws2_index)] = data2['leave_balance']
    ws2['E1' +str(ws2_index)] = data2['Number_of_days']
    ws2['F1' +str(ws2_index)] = data2['status']
    ws2['G1' +str(ws2_index)] = data2['comments']
    ws2_index += 1
wb.save(f'output/orange_report_{current_time.strftime('%d_%m_%Y_%H_%M')}.xlsx')
    
    
#for time
driver.find_element(By.XPATH, "/html/body/div/div[1]/div[1]/aside/nav/div[2]/ul/li[4]/a/span").click()
time.sleep(2)
elements3 = driver.find_elements(By.XPATH, '/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[3]/div/div[2]/div')
data_list3=[]

for i in range(1, len(elements3)+1):
    data3 = {}
    data3['Employee Name'] = driver.find_element(By.XPATH, '/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[3]/div/div[2]/div[' + str(i) + ']/div/div[1]/div').text
    data3['Timesheet period'] = driver.find_element(By.XPATH, '/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[3]/div/div[2]/div[' + str(i) + ']/div/div[1]/div').text
    data3['Action'] = driver.find_element(By.XPATH, '/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[3]/div/div[2]/div[' + str(i) + ']/div/div[1]/div').text
    data_list3.append(data3)
print(len(elements3))

ws3 = wb.create_sheet('time_orange')
ws3_index = ws3.max_row + 2

ws3['A1'] = "Employee Name"
ws3['B1'] = "Timesheet period"
ws3['C1'] = "Action"

for data in data_list3:
    
    ws3['A1' +str(ws3_index)] = data3['Employee Name']
    ws3['B1' +str(ws3_index)] = data3['Timesheet period']
    ws3['C1' +str(ws3_index)] = data3['Action']
    ws3_index += 1
wb.save(f'output/orange_report_{current_time.strftime('%d_%m_%Y_%H_%M')}.xlsx')

sys.exit()

# recruitment
driver.find_element(By.XPATH, '/html/body/div/div[1]/div[1]/aside/nav/div[2]/ul/li[5]/a/span').click()
time.sleep(2)
elements4i = driver.find_elements(By.XPATH, '/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[3]/div/div[2]/div')
data_list4i = []

for i in range(1, len(elements4i)+1):
    data4i = {}
    data4i["vacancy"] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[3]/div/div[2]/div[" + str(i) + "]/div/div[2]/div").text
    data4i["candidate"] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[3]/div/div[2]/div[" + str(i) + "]/div/div[3]/div").text
    data4i['hiring_manager'] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[3]/div/div[2]/div[" + str(i) + "]/div/div[4]/div").text
    data4i['date_of_application'] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[3]/div/div[2]/div[" + str(i) + "]/div/div[5]/div").text
    data4i['status'] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[3]/div/div[2]/div[" + str(i) + "]/div/div[6]/div").text
    data_list4i.append(data4i)
print(len(elements4i))

driver.find_element(By.XPATH, '/html/body/div/div[1]/div[2]/div[2]/div/div[2]/div[4]/nav/ul/li[2]/button').click()
time.sleep(5)

#my info

driver.find_element(By.XPATH, "/html/body/div/div[1]/div[1]/aside/nav/div[2]/ul/li[6]/a/span").click()
time.sleep(2)
data_list5 = []
data5 = {}
data5['file_name'] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div/div/div[2]/div[3]/div[3]/div/div[2]/div/div/div[2]").text
data5['Description'] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div/div/div[2]/div[3]/div[3]/div/div[2]/div/div/div[3]").text
data5['size'] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div/div/div[2]/div[3]/div[3]/div/div[2]/div/div/div[4]").text
data5['type'] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div/div/div[2]/div[3]/div[3]/div/div[2]/div/div/div[5]").text
data5['date_added'] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div/div/div[2]/div[3]/div[3]/div/div[2]/div/div/div[6]").text
data5['added_by'] = driver.find_element(By.XPATH, "/html/body/div/div[1]/div[2]/div[2]/div/div/div/div[2]/div[3]/div[3]/div/div[2]/div/div/div[7]").text
data_list5.append(data5)
print(data_list5)


#claim
driver.find_element(By.XPATH, '/html/body/div/div[1]/div[1]/aside/nav/div[2]/ul/li[11]/a/span').click()
time.sleep(2)
elements5 = driver.find_elements(By.XPATH, '/html/body/div/div[1]/div[2]/div[2]/div[2]/div[3]/div/div[2]/div')
time.sleep(2)
data_list6 = []

for i in range(1, len(elements5)+1):
    data6 = {}
    data['reference_id'] = driver.find_element("/html/body/div/div[1]/div[2]/div[2]/div[2]/div[3]/div/div[2]/div[" + str(i) + "]/div/div[1]/div").text
    data['Nom_de_employee'] = driver.find_element("/html/body/div/div[1]/div[2]/div[2]/div[2]/div[3]/div/div[2]/div[" + str(i) + "]/div/div[2]/div").text
    data['Event_Name'] = driver.find_element("/html/body/div/div[1]/div[2]/div[2]/div[2]/div[3]/div/div[2]/div[" + str(i) + "]/div/div[3]/div").text
    data['Description'] = driver.find_element("/html/body/div/div[1]/div[2]/div[2]/div[2]/div[3]/div/div[2]/div[" + str(i) + "]/div/div[4]/div").text
    data['Devise'] = driver.find_element("/html/body/div/div[1]/div[2]/div[2]/div[2]/div[3]/div/div[2]/div[" + str(i) + "]/div/div[5]/div").text
    data['submitted_data'] = driver.find_element("/html/body/div/div[1]/div[2]/div[2]/div[2]/div[3]/div/div[2]/div[" + str(i) + "]/div/div[6]/div").text
    data['status'] = driver.find_element("/html/body/div/div[1]/div[2]/div[2]/div[2]/div[3]/div/div[2]/div[" + str(i) + "]/div/div[7]/div").text
    data['Amount'] = driver.find_element("/html/body/div/div[1]/div[2]/div[2]/div[2]/div[3]/div/div[2]/div[" + str(i) + "]/div/div[8]/div").text
    data_list6.append(data6)
print(len(elements5))



    